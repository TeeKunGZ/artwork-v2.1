import asyncio
import io
import tempfile
import unittest
from pathlib import Path

import openpyxl
from fastapi import HTTPException, UploadFile
from PIL import Image

from app.routers import core
from app.services.excel_writer import generate_excel_bytes


def _png_bytes() -> bytes:
    buf = io.BytesIO()
    Image.new("RGBA", (8, 8), (255, 0, 0, 255)).save(buf, format="PNG")
    return buf.getvalue()


def _template_bytes() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["B2"] = "preserve-me"
    ws["D2"] = "AAAA-BBB-CCC-DDD"
    for col in ("AJ", "AN"):
        ws.column_dimensions[col].width = 16
    ws.row_dimensions[2].height = 80
    out = io.BytesIO()
    wb.save(out)
    wb.close()
    return out.getvalue()


class ExcelWriterTests(unittest.TestCase):
    def test_uploaded_template_is_preserved_and_aj_an_embed(self):
        records = [{
            "item_id": "AAAA-BBB-CCC-DDD",
            "style": "AAAA-DDD",
            "cw": "BBB",
            "org_code": "CCC",
            "team": "Team",
            "color": "Red",
        }]
        image = _png_bytes()
        crops = [("aj.png", image), ("an.png", image)]
        mappings = {
            "aj.png": {"item_id": "AAAA-BBB-CCC-DDD", "col": "AJ"},
            "an.png": {"item_id": "AAAA-BBB-CCC-DDD", "col": "AN"},
        }

        result = generate_excel_bytes(records, crops, mappings, _template_bytes())
        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active

        self.assertEqual(ws["B2"].value, "preserve-me")
        self.assertEqual(ws["D2"].value, "AAAA-BBB-CCC-DDD")
        self.assertEqual(len(ws._images), 2)
        wb.close()


class DatasetSafetyTests(unittest.TestCase):
    def test_save_dataset_rejects_path_traversal_label(self):
        async def run():
            with tempfile.TemporaryDirectory() as tmp:
                old_dataset, old_history = core.DATASET_DIR, core.HISTORY_DIR
                core.DATASET_DIR = str(Path(tmp) / "dataset")
                core.HISTORY_DIR = str(Path(tmp) / "history")
                try:
                    upload = UploadFile(file=io.BytesIO(_png_bytes()), filename="crop.png")
                    with self.assertRaises(HTTPException) as ctx:
                        await core.save_dataset(
                            file=upload,
                            label="../bad",
                            item_id="AAAA-BBB-CCC-DDD",
                            _={"emp_id": "tester"},
                        )
                    self.assertEqual(ctx.exception.status_code, 400)
                finally:
                    core.DATASET_DIR = old_dataset
                    core.HISTORY_DIR = old_history

        asyncio.run(run())

    def test_save_dataset_allows_aj_and_an(self):
        async def run():
            with tempfile.TemporaryDirectory() as tmp:
                old_dataset, old_history = core.DATASET_DIR, core.HISTORY_DIR
                core.DATASET_DIR = str(Path(tmp) / "dataset")
                core.HISTORY_DIR = str(Path(tmp) / "history")
                try:
                    for label in ("AJ", "AN"):
                        upload = UploadFile(file=io.BytesIO(_png_bytes()), filename="crop.png")
                        result = await core.save_dataset(
                            file=upload,
                            label=label,
                            item_id="AAAA-BBB-CCC-DDD",
                            _={"emp_id": "tester"},
                        )
                        self.assertEqual(result["status"], "success")
                        self.assertTrue((Path(core.HISTORY_DIR) / label / "AAAA-BBB-CCC-DDD.png").exists())
                finally:
                    core.DATASET_DIR = old_dataset
                    core.HISTORY_DIR = old_history

        asyncio.run(run())


if __name__ == "__main__":
    unittest.main()
