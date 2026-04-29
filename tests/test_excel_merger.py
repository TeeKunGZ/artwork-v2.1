import asyncio
import io
import unittest

import openpyxl
from fastapi import UploadFile
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill
from PIL import Image

from app.routers.export import _read_xlsx_upload, merge_excel
from app.services.excel_merger import merge_excel_bytes


def _png_bytes(color=(255, 0, 0, 255)) -> bytes:
    buf = io.BytesIO()
    Image.new("RGBA", (10, 10), color).save(buf, format="PNG")
    return buf.getvalue()


def _workbook_bytes(rows, image_rows=None) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["D1"] = "ITEM_ID"
    ws["E1"] = "CW"
    ws.column_dimensions["N"].width = 16
    image_rows = set(image_rows or [])

    for idx, item_id in enumerate(rows, start=2):
        ws.cell(row=idx, column=4, value=item_id)
        ws.cell(row=idx, column=5, value=f"CW{idx}")
        ws.cell(row=idx, column=5).font = Font(bold=True)
        ws.cell(row=idx, column=5).fill = PatternFill("solid", fgColor="FFFF00")
        ws.row_dimensions[idx].height = 55
        if idx in image_rows:
            img = XLImage(io.BytesIO(_png_bytes()))
            img.width = 20
            img.height = 20
            ws.add_image(img, f"N{idx}")

    out = io.BytesIO()
    wb.save(out)
    wb.close()
    return out.getvalue()


class ExcelMergerTests(unittest.TestCase):
    def test_appends_non_duplicate_rows_and_skips_duplicates(self):
        result, summary = merge_excel_bytes(
            _workbook_bytes(["AAAA-BBB-CCC-DDD"]),
            [("import.xlsx", _workbook_bytes(["AAAA-BBB-CCC-DDD", "EEEE-FFF-GGG-HHH"]))],
        )

        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active
        self.assertEqual(summary.added, 1)
        self.assertEqual(summary.skipped, 1)
        self.assertEqual(ws["D2"].value, "AAAA-BBB-CCC-DDD")
        self.assertEqual(ws["D3"].value, "EEEE-FFF-GGG-HHH")
        wb.close()

    def test_merges_multiple_import_files_in_order_and_skips_blank_ids(self):
        result, summary = merge_excel_bytes(
            _workbook_bytes(["BASE-BBB-CCC-DDD"]),
            [
                ("one.xlsx", _workbook_bytes(["ONEE-BBB-CCC-DDD", ""])),
                ("two.xlsx", _workbook_bytes(["TWOO-BBB-CCC-DDD"])),
            ],
        )

        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active
        self.assertEqual(summary.added, 2)
        self.assertEqual(summary.skipped, 1)
        self.assertEqual(ws["D3"].value, "ONEE-BBB-CCC-DDD")
        self.assertEqual(ws["D4"].value, "TWOO-BBB-CCC-DDD")
        wb.close()

    def test_copies_style_row_height_and_embedded_images(self):
        result, summary = merge_excel_bytes(
            _workbook_bytes(["BASE-BBB-CCC-DDD"]),
            [("import.xlsx", _workbook_bytes(["IMGE-BBB-CCC-DDD"], image_rows={2}))],
        )

        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active
        self.assertEqual(summary.added, 1)
        self.assertEqual(ws["D3"].value, "IMGE-BBB-CCC-DDD")
        self.assertTrue(ws["E3"].font.bold)
        self.assertEqual(ws["E3"].fill.fgColor.rgb, "00FFFF00")
        self.assertEqual(ws.row_dimensions[3].height, 55)
        self.assertEqual(len(ws._images), 1)
        self.assertEqual(ws._images[0].anchor._from.row, 2)
        self.assertEqual(ws._images[0].anchor._from.col, 13)
        wb.close()


class MergeUploadValidationTests(unittest.TestCase):
    def test_read_xlsx_upload_rejects_non_xlsx(self):
        async def run():
            upload = UploadFile(file=io.BytesIO(b"not excel"), filename="bad.txt")
            with self.assertRaises(ValueError):
                await _read_xlsx_upload(upload, "base_file")

        asyncio.run(run())

    def test_merge_endpoint_returns_json_for_missing_files(self):
        async def run():
            response = await merge_excel(base_file=None, import_files=[], _={"emp_id": "tester"})
            self.assertEqual(response.status_code, 400)

        asyncio.run(run())


if __name__ == "__main__":
    unittest.main()
