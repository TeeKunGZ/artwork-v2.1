"""
Excel & ZIP generation — column mapping verified จาก Templates.xlsx (2026-03)

Fixed text columns (เขียนโดยระบบ):
    C  (3)  = Style
    D  (4)  = Item id    ← row key
    E  (5)  = CW
    F  (6)  = ORG CODE
    G  (7)  = Team name
    M  (13) = Fabric col (color)

Image columns (embed รูปจาก crop):
    N, R, T, V, X, Z, AB, AD, AF, AH, AL, AP, AS, AV, AX, AZ
"""
from __future__ import annotations

import io
import os
import struct
import threading
import zipfile
from concurrent.futures import ThreadPoolExecutor

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils import column_index_from_string
from PIL import Image as PILImage

from app.services.columns import IMAGE_COL_LETTERS, normalize_image_col

TEMPLATE_PATH  = "Templates.xlsx"
FIRST_DATA_ROW = 2
_EMU_PER_PX    = 9525

# ── Template cache (avoid re-parsing XML on every request) ────────────────────
_template_cache_lock  = threading.Lock()
_template_cache_bytes: bytes | None = None
_template_cache_mtime: float = 0.0


def _get_template_bytes() -> bytes:
    """Return cached template bytes; reload only when file mtime changes."""
    global _template_cache_bytes, _template_cache_mtime
    mtime = os.path.getmtime(TEMPLATE_PATH)
    if _template_cache_bytes is not None and mtime == _template_cache_mtime:
        return _template_cache_bytes
    with _template_cache_lock:
        # double-check after acquiring lock
        mtime = os.path.getmtime(TEMPLATE_PATH)
        if _template_cache_bytes is not None and mtime == _template_cache_mtime:
            return _template_cache_bytes
        with open(TEMPLATE_PATH, "rb") as f:
            _template_cache_bytes = f.read()
        _template_cache_mtime = mtime
    return _template_cache_bytes


def _image_size_fast(data: bytes) -> tuple[int, int]:
    """Read image width/height from PNG IHDR header without full decode.
    Falls back to PIL for non-PNG formats."""
    # PNG signature (8 bytes) + IHDR chunk: length(4) + 'IHDR'(4) + width(4) + height(4)
    if len(data) >= 24 and data[:8] == b'\x89PNG\r\n\x1a\n':
        w, h = struct.unpack('>II', data[16:24])
        return w, h
    # Fallback for JPEG / other formats
    img = PILImage.open(io.BytesIO(data))
    size = img.size
    img.close()
    return size

_TEXT_COLS = {
    "style":    3,
    "item_id":  4,
    "cw":       5,
    "org_code": 6,
    "team":     7,
    "color":    13,
}

def _col_emu(sheet, col_letter: str) -> int:
    """Column width (character units) → EMU.  Calibri 11 / 96 DPI.
    รองรับ column_dimensions แบบ range (min/max) ที่ openpyxl เก็บ."""
    default_w = sheet.sheet_format.defaultColWidth or 8.43
    col_idx   = column_index_from_string(col_letter)

    # ค้นหาแบบ range-aware ก่อน (openpyxl อาจเก็บ width เป็น range เช่น min=18, max=54)
    w = None
    for dim in sheet.column_dimensions.values():
        if dim.min and dim.max and dim.width:
            if dim.min <= col_idx <= dim.max:
                w = dim.width
                break

    # fallback: ค้นหาด้วย key ปกติ
    if w is None:
        dim = sheet.column_dimensions.get(col_letter)
        w   = dim.width if (dim and dim.width) else default_w

    mdw = 7  # max-digit-width for Calibri 11 at 96 DPI
    px  = int(((256 * w + int(128 / mdw)) / 256) * mdw)
    return max(10 * _EMU_PER_PX, px * _EMU_PER_PX)


def _row_emu(sheet, row: int) -> int:
    """Row height (points) → EMU.  1 pt = 12700 EMU."""
    default_h = sheet.sheet_format.defaultRowHeight or 15.0
    dim = sheet.row_dimensions.get(row)
    h   = dim.height if (dim and dim.height) else default_h
    return max(12700, int(h * 12700))


def _build_row_map(sheet) -> dict[str, int]:
    row_map: dict[str, int] = {}
    id_col   = _TEXT_COLS["item_id"]
    scan_end = max(sheet.max_row, FIRST_DATA_ROW) + 1
    empty_streak = 0
    for r in range(FIRST_DATA_ROW, scan_end):
        val = sheet.cell(row=r, column=id_col).value
        if val:
            row_map[str(val).strip()] = r
            empty_streak = 0
        else:
            empty_streak += 1
            if empty_streak >= 50:
                break
    return row_map


def _next_empty_row(row_map: dict) -> int:
    if not row_map:
        return FIRST_DATA_ROW
    return max(row_map.values()) + 1


def generate_excel_bytes(
    records: list[dict],
    crops: list[tuple[str, bytes]],
    mappings: dict,
    template_bytes: bytes | None = None,
) -> bytes:
    if template_bytes is None and not os.path.exists(TEMPLATE_PATH):
        raise FileNotFoundError("ไม่พบไฟล์ Templates.xlsx")

    wb    = openpyxl.load_workbook(io.BytesIO(template_bytes or _get_template_bytes()))
    try:
        sheet = wb.active

        row_map  = _build_row_map(sheet)
        next_row = _next_empty_row(row_map)

        # ── Write text columns ───────────────────────────────────────────────────
        for rec in records:
            item_id = str(rec.get("item_id", "")).strip()
            if not item_id:
                continue

            if item_id in row_map:
                target_row = row_map[item_id]
            else:
                target_row       = next_row
                row_map[item_id] = target_row
                next_row        += 1

            sheet.cell(row=target_row, column=_TEXT_COLS["style"],    value=rec.get("style"))
            sheet.cell(row=target_row, column=_TEXT_COLS["item_id"],  value=item_id)
            sheet.cell(row=target_row, column=_TEXT_COLS["cw"],       value=rec.get("cw"))
            sheet.cell(row=target_row, column=_TEXT_COLS["org_code"], value=rec.get("org_code"))
            sheet.cell(row=target_row, column=_TEXT_COLS["team"],     value=rec.get("team"))
            sheet.cell(row=target_row, column=_TEXT_COLS["color"],    value=rec.get("color"))

        # ── Embed images ─────────────────────────────────────────────────────────
        # Pre-filter valid crops and pre-compute cell dimensions (thread-safe reads)
        valid_crops: list[tuple[str, bytes, str, int, int, int]] = []
        for fname, img_data in crops:
            if fname not in mappings:
                continue
            map_info          = mappings[fname]
            item_id           = str(map_info["item_id"]).strip()
            target_col_letter = normalize_image_col(map_info["col"])
            if item_id not in row_map:
                continue
            if target_col_letter not in IMAGE_COL_LETTERS:
                print(f"[ExcelWriter] Skipped col {target_col_letter}: ไม่อยู่ใน image column list")
                continue
            target_row = row_map[item_id]
            col_idx    = column_index_from_string(target_col_letter)
            cell_w_emu = _col_emu(sheet, target_col_letter)
            cell_h_emu = _row_emu(sheet, target_row)
            valid_crops.append((fname, img_data, target_col_letter, target_row, col_idx, cell_w_emu, cell_h_emu))

        # Prepare image objects in parallel (PIL size read + scaling calc)
        def _prepare_image(item):
            fname, img_data, target_col_letter, target_row, col_idx, cell_w_emu, cell_h_emu = item

            img_w_px, img_h_px = _image_size_fast(img_data)

            pad_emu  = _EMU_PER_PX * 2
            avail_w  = max(1, cell_w_emu - pad_emu * 2)
            avail_h  = max(1, cell_h_emu - pad_emu * 2)

            img_w_emu = img_w_px * _EMU_PER_PX
            img_h_emu = img_h_px * _EMU_PER_PX

            scale     = min(avail_w / img_w_emu, avail_h / img_h_emu)
            disp_w    = int(img_w_emu * scale)
            disp_h    = int(img_h_emu * scale)

            off_x = (cell_w_emu - disp_w) // 2
            off_y = (cell_h_emu - disp_h) // 2

            xl_img        = XLImage(io.BytesIO(img_data))
            xl_img.width  = max(1, disp_w // _EMU_PER_PX)
            xl_img.height = max(1, disp_h // _EMU_PER_PX)

            anchor = OneCellAnchor(
                _from=AnchorMarker(
                    col=col_idx - 1,  colOff=off_x,
                    row=target_row - 1, rowOff=off_y,
                ),
                ext=XDRPositiveSize2D(disp_w, disp_h),
            )
            xl_img.anchor = anchor

            print(f"[ExcelWriter] {fname} -> {target_col_letter}{target_row}: "
                  f"img={img_w_px}x{img_h_px}px "
                  f"cell={cell_w_emu}x{cell_h_emu}emu "
                  f"scale={scale:.3f} "
                  f"disp={disp_w}x{disp_h}emu "
                  f"off=({off_x},{off_y})")
            return xl_img

        # Use thread pool for parallel image prep; add_image must be sequential
        with ThreadPoolExecutor(max_workers=min(8, len(valid_crops) or 1)) as pool:
            prepared_images = list(pool.map(_prepare_image, valid_crops))

        for xl_img in prepared_images:
            sheet.add_image(xl_img)

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return out.getvalue()
    finally:
        wb.close()


def generate_zip_bytes(
    records: list[dict],
    crops: list[tuple[str, bytes]],
    mappings: dict,
) -> bytes:
    item_info = {r["item_id"]: r for r in records}
    out_zip   = io.BytesIO()

    with zipfile.ZipFile(out_zip, "w", zipfile.ZIP_DEFLATED) as zf:
        for fname, data in crops:
            if fname not in mappings:
                continue
            m        = mappings[fname]
            rec      = item_info.get(m["item_id"], {})
            team     = rec.get("team", "UNKNOWN").replace("/", "_").replace("\\", "_")
            safe_id  = m["item_id"].replace("/", "_").replace("\\", "_")
            col_name = m.get("col_name", "").replace("/", "_").replace("\\", "_")
            zf.writestr(f"{team}_{safe_id}_{m['col']}_{col_name}.png", data)

    out_zip.seek(0)
    return out_zip.getvalue()
