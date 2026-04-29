"""Excel workbook merge helpers for ArtPortal."""
from __future__ import annotations

import io
from copy import copy, deepcopy
from dataclasses import dataclass

import openpyxl
from openpyxl.drawing.image import Image as XLImage

ITEM_ID_COL = 4
FIRST_DATA_ROW = 2


@dataclass
class MergeSummary:
    added: int = 0
    skipped: int = 0
    errors: int = 0


def _item_id_at(sheet, row: int) -> str:
    return str(sheet.cell(row=row, column=ITEM_ID_COL).value or "").strip()


def _build_item_ids(sheet) -> set[str]:
    ids: set[str] = set()
    for row in range(FIRST_DATA_ROW, sheet.max_row + 1):
        item_id = _item_id_at(sheet, row)
        if item_id:
            ids.add(item_id)
    return ids


def _next_output_row(sheet) -> int:
    for row in range(sheet.max_row, FIRST_DATA_ROW - 1, -1):
        if any(sheet.cell(row=row, column=col).value not in (None, "") for col in range(1, sheet.max_column + 1)):
            return row + 1
    return FIRST_DATA_ROW


def _copy_row_cells(source_sheet, target_sheet, source_row: int, target_row: int) -> None:
    target_sheet.row_dimensions[target_row].height = source_sheet.row_dimensions[source_row].height
    for col in range(1, source_sheet.max_column + 1):
        src = source_sheet.cell(row=source_row, column=col)
        dst = target_sheet.cell(row=target_row, column=col)
        dst.value = src.value
        if src.has_style:
            dst.font = copy(src.font)
            dst.fill = copy(src.fill)
            dst.border = copy(src.border)
            dst.alignment = copy(src.alignment)
            dst.protection = copy(src.protection)
            dst.number_format = src.number_format
        if src.hyperlink:
            dst._hyperlink = copy(src.hyperlink)
        if src.comment:
            dst.comment = copy(src.comment)


def _anchor_row(anchor) -> int | None:
    marker = getattr(anchor, "_from", None)
    if marker is None:
        return None
    return int(marker.row) + 1


def _move_anchor_to_row(anchor, target_row: int):
    new_anchor = deepcopy(anchor)
    row_delta = (target_row - 1) - new_anchor._from.row
    new_anchor._from.row = target_row - 1
    to_marker = getattr(new_anchor, "to", None)
    if to_marker is not None:
        to_marker.row += row_delta
    return new_anchor


def _copy_row_images(source_sheet, target_sheet, source_row: int, target_row: int) -> int:
    copied = 0
    for image in getattr(source_sheet, "_images", []):
        if _anchor_row(image.anchor) != source_row:
            continue
        new_image = XLImage(io.BytesIO(image._data()))
        new_image.width = image.width
        new_image.height = image.height
        new_image.anchor = _move_anchor_to_row(image.anchor, target_row)
        target_sheet.add_image(new_image)
        copied += 1
    return copied


def merge_excel_bytes(base_bytes: bytes, import_files: list[tuple[str, bytes]]) -> tuple[bytes, MergeSummary]:
    """Merge import workbook first sheets into the base workbook first sheet.

    Duplicate ITEM_ID values in column D are skipped. Output keeps the base
    workbook as the template and appends non-duplicate import rows to it.
    """
    summary = MergeSummary()
    base_wb = openpyxl.load_workbook(io.BytesIO(base_bytes))
    import_workbooks = []
    try:
        base_sheet = base_wb.worksheets[0]
        existing_ids = _build_item_ids(base_sheet)
        next_row = _next_output_row(base_sheet)

        for _filename, data in import_files:
            import_wb = openpyxl.load_workbook(io.BytesIO(data))
            import_workbooks.append(import_wb)
            import_sheet = import_wb.worksheets[0]

            for source_row in range(FIRST_DATA_ROW, import_sheet.max_row + 1):
                item_id = _item_id_at(import_sheet, source_row)
                if not item_id:
                    summary.skipped += 1
                    continue
                if item_id in existing_ids:
                    summary.skipped += 1
                    continue

                try:
                    _copy_row_cells(import_sheet, base_sheet, source_row, next_row)
                    _copy_row_images(import_sheet, base_sheet, source_row, next_row)
                    existing_ids.add(item_id)
                    summary.added += 1
                    next_row += 1
                except Exception:
                    summary.errors += 1

        out = io.BytesIO()
        base_wb.save(out)
        out.seek(0)
        return out.getvalue(), summary
    finally:
        base_wb.close()
        for wb in import_workbooks:
            wb.close()
