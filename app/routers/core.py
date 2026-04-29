"""
Core processing routes:
  POST /api/extract-ai        — upload .ai file → artboards + text records
  POST /api/auto_detect       — OpenCV object detection on artboard image
  POST /api/save-crop-memory  — persist crop coordinates to DB
  GET  /api/get-template      — download Templates.xlsx
"""
from __future__ import annotations

import base64
import io
import json
import os
import re
from pathlib import Path

import cv2
import fitz
import numpy as np
import openpyxl
from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import FileResponse, JSONResponse
from PIL import Image as PILImage
from sqlalchemy.orm import Session

from app.db.base import get_db
from app.db.crud import get_all_crop_memory, upsert_crop_memory
from app.dependencies import get_current_user
from app.services.ocr_pipeline import extract_text_blocks
from app.services.parser import parse_item_records, _clean
from app.services.ai_status import get_registry, ModuleState
from app.config import settings
from app.services.columns import IMAGE_COL_LETTERS, normalize_image_col
import time
import uuid

router = APIRouter(tags=["Core"])

HISTORY_DIR = "history_db"
DATASET_DIR = "dataset"
MAX_UPLOAD_BYTES = settings.MAX_UPLOAD_MB * 1024 * 1024
ITEM_ID_SAFE_PATTERN = re.compile(r"^[A-Z0-9]+-[A-Z0-9]+-[A-Z0-9]+-[A-Z0-9]+$")


def _safe_child_dir(base_dir: str, child_name: str) -> Path:
    base = Path(base_dir).resolve()
    target = (base / child_name).resolve()
    if target != base and base not in target.parents:
        raise ValueError("Invalid path")
    return target


# ── Template download ─────────────────────────────────────────────────────────
@router.get("/get-template")
async def get_template():
    if os.path.exists("Templates.xlsx"):
        return FileResponse("Templates.xlsx", filename="Templates.xlsx")
    return JSONResponse(status_code=404, content={"message": "ไม่พบไฟล์"})


# ── Main AI extraction ────────────────────────────────────────────────────────
@router.post("/extract-ai")
async def extract_ai_image(
    file: UploadFile = File(...),
    excel_file: UploadFile = File(...),
    use_ai: str = Form("true"),
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    try:
        is_ai_enabled = use_ai.lower() == "true"

        # --- Check for duplicate item IDs already in the Excel template -------
        sheet = openpyxl.load_workbook(
            io.BytesIO(await excel_file.read()), data_only=True
        ).active
        existing = {
            str(sheet.cell(row=r, column=4).value).strip()
            for r in range(2, sheet.max_row + 1)
            if sheet.cell(row=r, column=4).value
        }

        ai_content = await file.read()

        # --- File size guard --------------------------------------------------
        if len(ai_content) > MAX_UPLOAD_BYTES:
            return JSONResponse(
                status_code=413,
                content={"status": "error", "message": f"ไฟล์ใหญ่เกิน {settings.MAX_UPLOAD_MB} MB"},
            )

        # --- OCR via fallback chain -------------------------------------------
        text_blocks = extract_text_blocks(ai_content, use_ai=is_ai_enabled)
        raw_text = "\n".join(b["text"] for b in text_blocks)

        # --- XMP metadata injection ------------------------------------------
        xmp = re.search(
            r"<\?xpacket begin.*?\?>(.*?)<\?xpacket end",
            ai_content[:131_072].decode("latin-1", errors="ignore"),
            re.DOTALL,
        )
        if xmp:
            for tag in ("dc:title", "dc:description", "xmp:Label", "photoshop:Headline"):
                m = re.search(
                    rf"<{re.escape(tag)}[^>]*>(.*?)</{re.escape(tag)}>",
                    xmp.group(1), re.DOTALL,
                )
                if m:
                    text_blocks.insert(0, {
                        "text": _clean(re.sub(r"<[^>]+>", "", m.group(1))),
                        "x0": 0, "y0": 0, "x1": 0, "y1": 0, "page": -1,
                        "source": "xmp",
                    })

        doc = fitz.open("pdf", ai_content)
        for i in range(len(doc)):
            raw_text += "\n" + doc.load_page(i).get_text("text")

        records, global_team, all_colors = parse_item_records(
            text_blocks, raw_text, ai_content
        )

        # --- Duplicate guard --------------------------------------------------
        dups = [r["item_id"] for r in records if r["item_id"] in existing]
        if dups:
            doc.close()
            return JSONResponse(
                status_code=400,
                content={
                    "status": "error",
                    "message": f"❌ พบ ITEM_ID ซ้ำใน Excel!\nรายการ: {', '.join(dups)}",
                },
            )

        # --- Auto-mapped history images --------------------------------------
        auto_mapped = []
        if os.path.isdir(HISTORY_DIR):
            for col_dir in os.listdir(HISTORY_DIR):
                col_dir = normalize_image_col(col_dir)
                if col_dir not in IMAGE_COL_LETTERS:
                    continue
                col_path = os.path.join(HISTORY_DIR, col_dir)
                if not os.path.isdir(col_path):
                    continue
                for rec in records:
                    safe_id = rec["item_id"].replace("/", "_").replace("\\", "_")
                    img_file = os.path.join(col_path, f"{safe_id}.png")
                    if os.path.exists(img_file):
                        with open(img_file, "rb") as f_img:
                            b64 = base64.b64encode(f_img.read()).decode()
                        auto_mapped.append({
                            "itemId": rec["item_id"],
                            "col": col_dir,
                            "imageBase64": f"data:image/png;base64,{b64}",
                        })

        # --- Render artboards ------------------------------------------------
        artboards = []
        for i in range(len(doc)):
            page = doc.load_page(i)
            pix = page.get_pixmap(matrix=fitz.Matrix(2.0, 2.0), alpha=True)
            img = PILImage.frombytes("RGBA", [pix.width, pix.height], pix.samples)
            bbox = img.getbbox() or (0, 0, pix.width, pix.height)
            buf = io.BytesIO()
            img.save(buf, format="PNG")
            artboards.append({
                "artboard_number": i + 1,
                "filename": f"Artboard{i + 1}.png",
                "image_base64": f"data:image/png;base64,{base64.b64encode(buf.getvalue()).decode()}",
                "width_cm": (pix.width / 144) * 2.54,
                "height_cm": (pix.height / 144) * 2.54,
                "auto_crop_box": {
                    "x": bbox[0], "y": bbox[1],
                    "width": bbox[2] - bbox[0],
                    "height": bbox[3] - bbox[1],
                },
            })
            buf.close()

        doc.close()

        return {
            "status": "success",
            "artboards": artboards,
            "text_data": records,
            "auto_mapped": auto_mapped,
            "team": global_team,
            "colors": all_colors,
            "crop_memory": get_all_crop_memory(db),
        }

    except Exception as exc:
        import traceback
        traceback.print_exc()
        return JSONResponse(
            status_code=400,
            content={"status": "error", "message": str(exc)},
        )


# ── Object detection (GrabCut + Contour fallback) ─────────────────────────────
@router.post("/auto_detect")
async def auto_detect_objects(
    file: UploadFile = File(...),
    _: dict = Depends(get_current_user),
):
    reg   = get_registry()
    start = time.time()
    reg.update("cv2_detect", ModuleState.RUNNING, "กำลังสแกนหาชิ้นส่วน...")

    try:
        contents = await file.read()
        if len(contents) > MAX_UPLOAD_BYTES:
            raise HTTPException(
                status_code=413,
                detail=f"File too large. Maximum size is {settings.MAX_UPLOAD_MB} MB",
            )
        nparr    = np.frombuffer(contents, np.uint8)
        img_orig = cv2.imdecode(nparr, cv2.IMREAD_UNCHANGED)
        if img_orig is None:
            raise HTTPException(status_code=400, detail="Unable to read image file")
        h_img, w_img = img_orig.shape[:2]

        # ── แปลงเป็น BGR เสมอ สำหรับ GrabCut ──────────────────────────────
        if img_orig.ndim == 2:
            img_bgr = cv2.cvtColor(img_orig, cv2.COLOR_GRAY2BGR)
        elif img_orig.shape[2] == 4:
            img_bgr = cv2.cvtColor(img_orig, cv2.COLOR_BGRA2BGR)
        else:
            img_bgr = img_orig.copy()

        # ── GrabCut: ตัด border 5% ออก เพื่อไม่ให้ frame กลายเป็น foreground ─
        margin_x = int(w_img * 0.05)
        margin_y = int(h_img * 0.05)
        rect     = (margin_x, margin_y,
                    w_img - margin_x * 2,
                    h_img - margin_y * 2)

        mask_gc  = np.zeros((h_img, w_img), np.uint8)
        bgd_model = np.zeros((1, 65), np.float64)
        fgd_model = np.zeros((1, 65), np.float64)

        try:
            cv2.grabCut(img_bgr, mask_gc, rect, bgd_model, fgd_model,
                        iterCount=3, mode=cv2.GC_INIT_WITH_RECT)
            # pixels ที่ GrabCut ตัดสินว่าเป็น foreground (probable หรือ definite)
            fg_mask = np.where((mask_gc == cv2.GC_FGD) | (mask_gc == cv2.GC_PR_FGD),
                               255, 0).astype(np.uint8)
        except Exception:
            # GrabCut fail → fallback ใช้ alpha หรือ Canny เหมือนเดิม
            if img_orig.ndim == 3 and img_orig.shape[2] == 4:
                fg_mask = img_orig[:, :, 3]
            else:
                gray    = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)
                fg_mask = cv2.Canny(cv2.GaussianBlur(gray, (5, 5), 0), 30, 150)

        # ── Morphology: เชื่อม pixels ที่อยู่ใกล้กันในโลโก้เดียวกัน ──────────
        kernel  = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (15, 15))
        fg_mask = cv2.morphologyEx(fg_mask, cv2.MORPH_CLOSE, kernel, iterations=2)
        fg_mask = cv2.dilate(fg_mask, kernel, iterations=1)

        # ── Connected components → bounding boxes ────────────────────────────
        n_labels, _, stats, _ = cv2.connectedComponentsWithStats(fg_mask, connectivity=8)

        min_area = (w_img * h_img) * 0.002   # อย่างน้อย 0.2% ของภาพ
        max_area = (w_img * h_img) * 0.85    # ไม่เกิน 85% (ตัด background ออก)

        detected = []
        for i in range(1, n_labels):          # label 0 = background
            x = int(stats[i, cv2.CC_STAT_LEFT])
            y = int(stats[i, cv2.CC_STAT_TOP])
            w = int(stats[i, cv2.CC_STAT_WIDTH])
            h = int(stats[i, cv2.CC_STAT_HEIGHT])
            area = int(stats[i, cv2.CC_STAT_AREA])
            if min_area <= area <= max_area and w > 20 and h > 20:
                detected.append({"x": x, "y": y, "w": w, "h": h})

        # ── NMS: ลบ box ที่ซ้อนทับกันเกิน 60% ───────────────────────────────
        def _nms(boxes: list, iou_thresh: float = 0.6) -> list:
            if not boxes: return []
            b      = np.array([[d["x"], d["y"], d["x"]+d["w"], d["y"]+d["h"]] for d in boxes], float)
            areas  = (b[:,2]-b[:,0]) * (b[:,3]-b[:,1])
            order  = areas.argsort()[::-1]
            keep   = []
            while order.size:
                i = order[0]; keep.append(i)
                xx1 = np.maximum(b[i,0], b[order[1:],0]); yy1 = np.maximum(b[i,1], b[order[1:],1])
                xx2 = np.minimum(b[i,2], b[order[1:],2]); yy2 = np.minimum(b[i,3], b[order[1:],3])
                inter = np.maximum(0, xx2-xx1) * np.maximum(0, yy2-yy1)
                iou   = inter / (areas[i] + areas[order[1:]] - inter + 1e-6)
                order = order[1:][iou < iou_thresh]
            return [detected[k] for k in keep]

        result  = _nms(detected)
        elapsed = int((time.time() - start) * 1000)
        reg.update("cv2_detect", ModuleState.IDLE,
                   f"เจอ {len(result)} ชิ้นส่วน ({elapsed}ms)", duration_ms=elapsed)

        return {"status": "success", "objects": result}

    except HTTPException:
        elapsed = int((time.time() - start) * 1000)
        reg.update("cv2_detect", ModuleState.ERROR, "Invalid image input", duration_ms=elapsed)
        raise
    except Exception as exc:
        elapsed = int((time.time() - start) * 1000)
        reg.update("cv2_detect", ModuleState.ERROR, str(exc), duration_ms=elapsed)
        raise HTTPException(status_code=400, detail=str(exc))


# ── Crop memory ───────────────────────────────────────────────────────────────
@router.post("/save-crop-memory")
async def save_crop_memory(
    item_id: str = Form(...),
    column: str = Form(...),
    crop_data: str = Form(...),
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    upsert_crop_memory(db, item_id, column, json.loads(crop_data), current_user["emp_id"])
    return {"status": "success"}


# ── Dataset saving (used by app.js after each crop) ───────────────────────────
@router.post("/save-dataset")
async def save_dataset(
    file: UploadFile = File(...),
    label: str = Form(""),
    item_id: str = Form(""),
    _: dict = Depends(get_current_user),
):
    try:
        content = await file.read()
        if len(content) > MAX_UPLOAD_BYTES:
            raise HTTPException(
                status_code=413,
                detail=f"File too large. Maximum size is {settings.MAX_UPLOAD_MB} MB",
            )

        label = normalize_image_col(label)
        if label not in IMAGE_COL_LETTERS:
            raise HTTPException(status_code=400, detail="Invalid dataset label")

        item_id = (item_id or "").strip().upper()
        if item_id and not ITEM_ID_SAFE_PATTERN.fullmatch(item_id):
            raise HTTPException(status_code=400, detail="Invalid ITEM ID")

        # Persist to dataset/<label>/ for AI training
        class_dir = _safe_child_dir(DATASET_DIR, label)
        class_dir.mkdir(parents=True, exist_ok=True)
        fname = f"crop_{int(time.time() * 1000)}_{uuid.uuid4().hex[:8]}.png"
        with open(class_dir / fname, "wb") as f:
            f.write(content)

        # Persist to history_db/<label>/<item_id>.png for auto-mapping on re-upload
        if item_id:
            hist_dir = _safe_child_dir(HISTORY_DIR, label)
            hist_dir.mkdir(parents=True, exist_ok=True)
            with open(hist_dir / f"{item_id}.png", "wb") as f:
                f.write(content)

        return {"status": "success"}

    except HTTPException:
        raise
    except Exception as exc:
        return JSONResponse(
            status_code=400,
            content={"status": "error", "message": str(exc)},
        )


# ── Template column info ──────────────────────────────────────────────────────
